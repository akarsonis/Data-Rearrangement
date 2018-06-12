import xlrd
import openpyxl
from openpyxl import Workbook

wb = Workbook()
ws = wb.active

workbook = xlrd.open_workbook('CSVTest.xlsx')
worksheet = workbook.sheet_by_index(0)

def good_shoe():
    global a
    a = str(a)
    
    a = a.replace('"', '')
    a = a.replace('}', '')
    a = a.replace('{', '')
    a = a.replace('[', '')
    a = a.replace(']', '')
    a = a.split(',') 
    
    for i in range(len(a) - 1, -1, -1):
        a[i] = a[i].replace('CardsList:CardName:', '')
        a[i] = a[i].replace('CardName:', '')
        if 'Deck' in a[i] or 'timeStamp' in a[i] or 'Cards' in a[i]:
            del a[i]
            
    for i in range(len(a) - 1, -1, -1):
        a[i] = a[i].replace('h', '')
        a[i] = a[i].replace('c', '')
        a[i] = a[i].replace('d', '')
        a[i] = a[i].replace('s', '')
        a[i] = a[i].replace('rounI:', '')


c = -1
for row in range(4, worksheet.nrows):
    a = worksheet.cell_value(row, 6)
    good_shoe()
    #steps
    c +=3
    r = 3
    for p in a:
        if p != 'J' and p != 'Q' and p != 'K' and p != 'A' and p != 'T':
            ws.cell(row=r, column=c).value = int(p)
        else:
            ws.cell(row=r, column=c).value = str(p)
        r += 1

time_start = 2
for row in range(4, worksheet.nrows):
    a = worksheet.cell_value(row, 3)
    sfuffle_start = xlrd.xldate_as_tuple(a, workbook.datemode)
    sfuffle_start = list(sfuffle_start)
    sfuffle_start[0], sfuffle_start[1], sfuffle_start[2] = sfuffle_start[2], sfuffle_start[1], sfuffle_start[0]
    del sfuffle_start[5]
    sfuffle_start = str(sfuffle_start)
    sfuffle_start = sfuffle_start.replace('[', '')
    sfuffle_start = sfuffle_start.replace(']', '')
    sfuffle_start = sfuffle_start[:12].replace(',', '.') + sfuffle_start[12:]
    sfuffle_start = sfuffle_start[:12] + sfuffle_start[12:].replace(',', ':')
    sfuffle_start = sfuffle_start[:10] + sfuffle_start[10:14].replace('.', ' - ') + sfuffle_start[14:]
    sfuffle_start = sfuffle_start.replace(' ', '')
    sfuffle_start = sfuffle_start[sfuffle_start.find('-'):]
    sfuffle_start = sfuffle_start.replace('-', '')    
    ws.cell(row= 2, column= time_start).value = sfuffle_start
    #steps
    time_start += 3
    
time_end = 3
for row in range(4, worksheet.nrows):
    b = worksheet.cell_value(row, 4)
    sfuffle_end = xlrd.xldate_as_tuple(b, workbook.datemode)
    sfuffle_end = list(sfuffle_end)
    sfuffle_end[0], sfuffle_end[1], sfuffle_end[2] = sfuffle_end[2], sfuffle_end[1], sfuffle_end[0]
    del sfuffle_end[5]
    sfuffle_end = str(sfuffle_end)
    sfuffle_end = sfuffle_end.replace('[', '')
    sfuffle_end = sfuffle_end.replace(']', '')
    sfuffle_end = sfuffle_end[:12].replace(',', '.') + sfuffle_end[12:]
    sfuffle_end = sfuffle_end[:12] + sfuffle_end[12:].replace(',', ':')
    sfuffle_end = sfuffle_end[:10] + sfuffle_end[10:14].replace('.', ' - ') + sfuffle_end[14:]
    sfuffle_end = sfuffle_end.replace(' ', '')
    sfuffle_end = sfuffle_end[sfuffle_end.find('-'):]
    sfuffle_end = sfuffle_end.replace('-', '')    
    ws.cell(row= 2, column= time_end).value = sfuffle_end
    #steps
    time_end += 3
    
shuffle_date_count = 2
for row in range(4, worksheet.nrows):
    d = worksheet.cell_value(row, 4)
    shuffle_date = xlrd.xldate_as_tuple(d, workbook.datemode)
    shuffle_date = list(shuffle_date)
    shuffle_date = shuffle_date[2], shuffle_date[1], shuffle_date[0]
    shuffle_date = str(shuffle_date)
    shuffle_date = shuffle_date.replace('[', '')
    shuffle_date = shuffle_date.replace(']', '')
    shuffle_date = shuffle_date.replace('(', '')
    shuffle_date = shuffle_date.replace(')', '')
    shuffle_date = shuffle_date.replace(' ', '')
    shuffle_date = shuffle_date.replace(',', '.')
    ws.cell(row= 1, column= shuffle_date_count).value = shuffle_date
    #steps
    shuffle_date_count += 3
    
wb.save('Nice Shoes.xlsx')